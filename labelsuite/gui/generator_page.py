"""목록 생성 탭 — 레거시 LiGen의 tkinter UI를 대체.

개선점: 백그라운드 로드/생성(프리즈 없음), 연/월 일괄 선택, 행 이슈 패널,
생성 즉시 검사 탭으로 메모리 핸드오프(list_generated 시그널).
"""

from __future__ import annotations

import os
from datetime import date

from PySide6.QtCore import QAbstractTableModel, QModelIndex, Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTableView,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from labelsuite.core.config import AppConfig
from labelsuite.core.list_generator import (
    ColumnMaps,
    GenerationResult,
    extract_available_dates,
)
from labelsuite.core.schema import (
    CANONICAL_COLUMNS,
    STANDARD_COLUMN,
    LabelRecord,
    save_inspection_list,
)
from labelsuite.gui.widgets.date_tree import DateTreeWidget
from labelsuite.gui.workers import ExcelLoadWorker, GenerateWorker

_FILE_LABELS = {
    "schedule": "주문일정 체크리스트",
    "product": "제품 품목번호 리스트",
    "bsc": "BSC FGD 리스트",
}


class RecordsTableModel(QAbstractTableModel):
    HEADERS = list(CANONICAL_COLUMNS) + [STANDARD_COLUMN]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._records: list[LabelRecord] = []

    def set_records(self, records: list[LabelRecord]) -> None:
        self.beginResetModel()
        self._records = list(records)
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self._records)

    def columnCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self.HEADERS)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or role != Qt.ItemDataRole.DisplayRole:
            return None
        record = self._records[index.row()]
        column = self.HEADERS[index.column()]
        return record.standard or "" if column == STANDARD_COLUMN else record.field(column)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        return self.HEADERS[section] if orientation == Qt.Orientation.Horizontal else section + 1


class GeneratorPage(QWidget):
    list_generated = Signal(list)   # list[LabelRecord] — 검사 탭 핸드오프
    status_message = Signal(str)

    def __init__(self, config: AppConfig, parent=None):
        super().__init__(parent)
        self.setObjectName("page")
        self.setAcceptDrops(True)
        self.config = config
        self.colmaps = ColumnMaps.from_config(config.column_maps_raw)
        self.frames: dict[str, object] = {"schedule": None, "product": None, "bsc": None}
        self._load_workers: dict[str, ExcelLoadWorker] = {}
        self._generate_worker: GenerateWorker | None = None
        self._result: GenerationResult | None = None
        self._build_ui()
        self._restore_last_files()

    # ---------- 드래그앤드롭: 시트명으로 파일 종류 자동 판별 ----------

    def dragEnterEvent(self, event) -> None:
        if any(url.toLocalFile().lower().endswith((".xlsx", ".xls"))
               for url in event.mimeData().urls()):
            event.acceptProposedAction()

    def dropEvent(self, event) -> None:
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if not path.lower().endswith((".xlsx", ".xls")):
                continue
            key = self._identify_file_key(path)
            if key is None:
                QMessageBox.warning(
                    self, "파일 판별 실패",
                    f"{os.path.basename(path)}\n"
                    "필요한 시트를 찾지 못했습니다. 버튼으로 직접 선택해 주세요.\n"
                    f"(기대 시트: {self.colmaps.schedule.sheet} / "
                    f"{self.colmaps.product.sheet} / {self.colmaps.bsc.sheet})")
                continue
            self._load_file(key, path, silent=False)
        event.acceptProposedAction()

    def _identify_file_key(self, path: str) -> str | None:
        """엑셀의 시트명을 읽어 어느 입력 파일인지 판별한다."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True)
            sheets = set(wb.sheetnames)
            wb.close()
        except Exception:
            sheets = set()
        for key in ("schedule", "product", "bsc"):
            if getattr(self.colmaps, key).sheet in sheets:
                return key
        # 폴백: 파일명 키워드
        name = os.path.basename(path)
        if "주문일정" in name or "일정" in name:
            return "schedule"
        if "품목" in name:
            return "product"
        if "BSC" in name.upper() or "FGD" in name.upper():
            return "bsc"
        return None

    def _build_ui(self) -> None:
        splitter = QSplitter(Qt.Orientation.Horizontal, self)

        # 좌측: 파일 선택 + 날짜 트리 + 실행 버튼
        left = QWidget()
        left_layout = QVBoxLayout(left)

        files_group = QGroupBox("입력 파일")
        files_layout = QVBoxLayout(files_group)
        files_layout.setSpacing(8)
        hint = QLabel("엑셀 파일을 이 창에 끌어다 놓으면 자동으로 분류됩니다")
        hint.setStyleSheet("color: #6b7280; font-weight: 400;")
        files_layout.addWidget(hint)
        self.file_status: dict[str, QLabel] = {}
        for key, label in _FILE_LABELS.items():
            row = QHBoxLayout()
            button = QPushButton(label)
            button.setMinimumWidth(170)
            button.clicked.connect(lambda _=False, k=key: self._pick_file(k))
            status = QLabel("선택되지 않음")
            status.setStyleSheet("color: #6b7280;")
            status.setWordWrap(True)
            self.file_status[key] = status
            row.addWidget(button)
            row.addWidget(status, stretch=1)
            files_layout.addLayout(row)
        left_layout.addWidget(files_group)

        self.date_tree = DateTreeWidget()
        self.date_tree.selection_changed.connect(self._update_buttons)
        left_layout.addWidget(self.date_tree, stretch=1)

        actions = QHBoxLayout()
        self.generate_button = QPushButton("리스트 생성")
        self.generate_button.setProperty("accent", True)
        self.generate_button.setEnabled(False)
        self.generate_button.clicked.connect(self._generate)
        self.save_button = QPushButton("엑셀로 저장")
        self.save_button.setEnabled(False)
        self.save_button.clicked.connect(self._save_xlsx)
        self.inspect_button = QPushButton("검사 탭으로 보내기 →")
        self.inspect_button.setProperty("accent", True)
        self.inspect_button.setEnabled(False)
        self.inspect_button.clicked.connect(self._send_to_inspector)
        actions.addWidget(self.generate_button)
        actions.addWidget(self.save_button)
        actions.addWidget(self.inspect_button)
        left_layout.addLayout(actions)

        # 우측: 미리보기 + 이슈
        right = QWidget()
        right_layout = QVBoxLayout(right)
        preview_group = QGroupBox("미리보기")
        preview_layout = QVBoxLayout(preview_group)
        self.table_model = RecordsTableModel(self)
        self.table = QTableView()
        self.table.setModel(self.table_model)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        preview_layout.addWidget(self.table)
        right_layout.addWidget(preview_group, stretch=3)

        issues_group = QGroupBox("경고/오류")
        issues_layout = QVBoxLayout(issues_group)
        self.issues_view = QTextEdit()
        self.issues_view.setReadOnly(True)
        issues_layout.addWidget(self.issues_view)
        right_layout.addWidget(issues_group, stretch=1)

        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.addWidget(splitter)

    def apply_config(self) -> None:
        """설정 변경 반영 — 컬럼 매핑을 다시 읽는다 (다음 로드부터 적용)."""
        self.colmaps = ColumnMaps.from_config(self.config.column_maps_raw)

    # ---------- 파일 로드 ----------

    def _restore_last_files(self) -> None:
        for key, path in self.config.settings.get("last_files", {}).items():
            if path and os.path.exists(path):
                self._load_file(key, path, silent=True)
            elif path:
                self.file_status[key].setText(f"이전 경로 없음: {os.path.basename(path)}")
                self.file_status[key].setStyleSheet("color: #b36b00;")

    def _pick_file(self, key: str) -> None:
        start_dir = os.path.dirname(self.config.settings["last_files"].get(key, "") or "")
        path, _ = QFileDialog.getOpenFileName(
            self, f"{_FILE_LABELS[key]} 선택", start_dir,
            "Excel 파일 (*.xlsx *.xls)")
        if path:
            self._load_file(key, path, silent=False)

    def _load_file(self, key: str, path: str, silent: bool) -> None:
        self.file_status[key].setText("읽는 중…")
        self.file_status[key].setStyleSheet("color: #0055aa;")
        worker = ExcelLoadWorker(key, path, self.colmaps, self)
        worker.finished_ok.connect(
            lambda k, frame, p=path, s=silent: self._on_file_loaded(k, frame, p, s))
        worker.failed.connect(self._on_file_failed)
        self._load_workers[key] = worker
        worker.start()

    def _on_file_loaded(self, key: str, frame, path: str, silent: bool) -> None:
        self.frames[key] = frame
        self.file_status[key].setText("✓ " + os.path.basename(path))
        self.file_status[key].setStyleSheet("color: #16a34a; font-weight: 600;")
        self.config.settings["last_files"][key] = path
        self.config.save_settings()
        if key == "schedule":
            # 스케줄 교체 시 날짜 트리·선택 초기화 (레거시 잔존 선택 버그 수정)
            self.date_tree.set_dates(extract_available_dates(frame, self.colmaps))
        self._update_buttons()
        if not silent:
            self.status_message.emit(f"{_FILE_LABELS[key]} 로드 완료: {os.path.basename(path)}")

    def _on_file_failed(self, key: str, message: str) -> None:
        self.frames[key] = None
        self.file_status[key].setText("로드 실패")
        self.file_status[key].setStyleSheet("color: red;")
        QMessageBox.warning(self, "파일 오류", f"{_FILE_LABELS[key]}\n{message}")
        self._update_buttons()

    # ---------- 생성 ----------

    def _update_buttons(self) -> None:
        ready = self.frames["schedule"] is not None and bool(self.date_tree.checked_dates())
        self.generate_button.setEnabled(ready)

    def _generate(self) -> None:
        selected: set[date] = self.date_tree.checked_dates()
        if self.frames["product"] is None and self.frames["bsc"] is None:
            answer = QMessageBox.question(
                self, "확인",
                "품목번호/BSC 리스트가 없어 GTIN·REF를 조회할 수 없습니다.\n"
                "그래도 생성할까요?")
            if answer != QMessageBox.StandardButton.Yes:
                return
        self.generate_button.setEnabled(False)
        self.status_message.emit("리스트 생성 중…")
        self._generate_worker = GenerateWorker(
            self.frames["schedule"], self.frames["product"], self.frames["bsc"],
            selected, self.colmaps,
            self.config.settings.get("country_standard_map", {}),
            int(self.config.settings.get("shelf_life_months", 36)), self)
        self._generate_worker.finished_ok.connect(self._on_generated)
        self._generate_worker.failed.connect(self._on_generate_failed)
        self._generate_worker.start()

    def _on_generated(self, result: GenerationResult) -> None:
        self._result = result
        self.table_model.set_records(result.records)
        self.table.resizeColumnsToContents()
        lines = [f"[{i.severity}] {i.row_index}행 {i.lot}: {i.message}"
                 for i in result.issues]
        self.issues_view.setPlainText("\n".join(lines) if lines else "이슈 없음")
        self.save_button.setEnabled(bool(result.records))
        self.inspect_button.setEnabled(bool(result.records))
        self._update_buttons()
        summary = (f"{len(result.records)}건 생성"
                   f" (경고 {result.warning_count}, 오류 {result.error_count})")
        self.status_message.emit(summary)
        if result.error_count or result.warning_count:
            QMessageBox.information(self, "생성 완료",
                                    summary + "\n자세한 내용은 경고/오류 패널을 확인하세요.")

    def _on_generate_failed(self, message: str) -> None:
        self._update_buttons()
        QMessageBox.critical(self, "오류", message)

    # ---------- 출력 ----------

    def _default_filename(self) -> str:
        dates = sorted(self._result.selected_dates) if self._result else []
        tag = ",".join(d.strftime("%y%m%d") for d in dates[:3])
        return f"Label Inspection List_{tag}.xlsx"

    def _save_xlsx(self) -> None:
        if not self._result or not self._result.records:
            return
        start = os.path.join(
            self.config.settings.get("save_directory") or os.path.expanduser("~"),
            self._default_filename())
        path, _ = QFileDialog.getSaveFileName(self, "검사 목록 저장", start,
                                              "Excel 파일 (*.xlsx)")
        if not path:
            return
        try:
            save_inspection_list(self._result.records, path)
        except OSError as exc:
            QMessageBox.critical(self, "저장 실패", str(exc))
            return
        self.config.settings["last_list_path"] = path
        self.config.save_settings()
        self.status_message.emit(f"저장 완료: {path}")

    def _send_to_inspector(self) -> None:
        if self._result and self._result.records:
            self.list_generated.emit(list(self._result.records))
