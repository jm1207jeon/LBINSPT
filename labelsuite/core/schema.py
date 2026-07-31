"""검사 목록의 표준 스키마 — 두 레거시 프로그램에 중복돼 있던 컬럼 계약의 단일 원천.

목록 파일 계약:
- 시트명 'Label Inspection List'
- 필수 컬럼 7개(CANONICAL_COLUMNS) + 선택적 'STANDARD' 컬럼(검사 규격 자동 선택용)
- GTIN은 항상 GTIN-14(14자리, 선행 0 유지)로 저장
"""

from __future__ import annotations

import re
from dataclasses import dataclass, replace
from typing import Iterable

from openpyxl import Workbook, load_workbook

CANONICAL_COLUMNS = ("LOT", "PRODUCTS", "PN", "REF", "MFG DATE", "EXP DATE", "GTIN")
STANDARD_COLUMN = "STANDARD"
SHEET_NAME = "Label Inspection List"

# 날짜 표기 규격: BSC(일본)는 점 구분, 그 외는 하이픈 구분
DATE_FORMAT_DEFAULT = "%Y-%m-%d"
DATE_FORMAT_BY_STANDARD = {"BSC": "%Y.%m.%d"}


class SchemaError(ValueError):
    """목록 파일이 스키마 계약을 위반했을 때."""


@dataclass(frozen=True)
class LabelRecord:
    """검사 목록의 한 행."""

    lot: str
    products: str
    pn: str
    ref: str
    mfg_date: str
    exp_date: str
    gtin: str
    standard: str | None = None

    def as_row(self, include_standard: bool) -> list[str]:
        row = [self.lot, self.products, self.pn, self.ref,
               self.mfg_date, self.exp_date, self.gtin]
        if include_standard:
            row.append(self.standard or "")
        return row

    def field(self, column: str) -> str:
        return {
            "LOT": self.lot, "PRODUCTS": self.products, "PN": self.pn,
            "REF": self.ref, "MFG DATE": self.mfg_date,
            "EXP DATE": self.exp_date, "GTIN": self.gtin,
        }[column]


def normalize_gtin14(raw: object) -> str:
    """GTIN 정규화의 유일한 규칙: 숫자만 남겨 14자리로 zero-pad.

    엑셀에서 흘러들어오는 float 흔적('8806173612345.0'), int, 문자열을 모두 수용한다.
    빈 값은 빈 문자열 그대로, 14자리 초과는 SchemaError.
    """
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text or text.lower() in ("nan", "none"):
        return ""
    # 엑셀 float 흔적 제거: '123.0' → '123' (지수 표기 포함)
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]
    elif re.fullmatch(r"[\d.]+[eE]\+?\d+", text):
        text = format(int(float(text)), "d")
    # GS1 AI(01) 접두가 붙은 바코드 텍스트 수용: '(01)08806...' / '0108806...'(16자리)
    match = re.fullmatch(r"\(01\)(\d{1,14})", text) or re.fullmatch(r"01(\d{14})", text)
    if match:
        text = match.group(1)
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    if len(digits) > 14:
        raise SchemaError(f"GTIN이 14자리를 초과합니다: {raw!r}")
    return digits.zfill(14)


def date_format_for(standard: str | None) -> str:
    """규격별 날짜 포맷. standards.json의 date_format이 우선하며 이는 폴백 규칙이다."""
    if standard and standard in DATE_FORMAT_BY_STANDARD:
        return DATE_FORMAT_BY_STANDARD[standard]
    return DATE_FORMAT_DEFAULT


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("nan", "none") else text


def load_inspection_list(path: str) -> tuple[list[LabelRecord], list[str]]:
    """7컬럼(레거시) 또는 8컬럼 목록 파일을 읽는다.

    반환: (records, warnings). 헤더가 계약과 다르면 SchemaError.
    레거시 파일의 13자리 GTIN은 로드 시 GTIN-14로 정규화된다.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise SchemaError("목록 파일이 비어 있습니다.")
        header = [_clean_cell(h) for h in header if h is not None]

        missing = [c for c in CANONICAL_COLUMNS if c not in header]
        if missing:
            raise SchemaError(
                "목록 파일의 컬럼이 올바르지 않습니다. "
                f"누락: {', '.join(missing)} / 발견된 헤더: {header}"
            )
        col_idx = {name: header.index(name) for name in CANONICAL_COLUMNS}
        has_standard = STANDARD_COLUMN in header
        if has_standard:
            col_idx[STANDARD_COLUMN] = header.index(STANDARD_COLUMN)

        records: list[LabelRecord] = []
        warnings: list[str] = []
        for i, row in enumerate(rows, start=2):
            if row is None or all(v is None or _clean_cell(v) == "" for v in row):
                continue

            def cell(name: str) -> str:
                idx = col_idx[name]
                return _clean_cell(row[idx]) if idx < len(row) else ""

            try:
                gtin = normalize_gtin14(cell("GTIN"))
            except SchemaError as exc:
                warnings.append(f"{i}행: {exc} — 원본 값을 유지합니다.")
                gtin = cell("GTIN")
            records.append(LabelRecord(
                lot=cell("LOT"), products=cell("PRODUCTS"), pn=cell("PN"),
                ref=cell("REF"), mfg_date=cell("MFG DATE"),
                exp_date=cell("EXP DATE"), gtin=gtin,
                standard=(cell(STANDARD_COLUMN) or None) if has_standard else None,
            ))
        if not records:
            warnings.append("목록에 데이터 행이 없습니다.")
        return records, warnings
    finally:
        wb.close()


def save_inspection_list(records: Iterable[LabelRecord], path: str,
                         include_standard: bool = True) -> None:
    """openpyxl로 저장. GTIN 셀은 텍스트 서식으로 선행 0을 보존한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    header = list(CANONICAL_COLUMNS)
    if include_standard:
        header.append(STANDARD_COLUMN)
    ws.append(header)
    gtin_col = CANONICAL_COLUMNS.index("GTIN") + 1
    for record in records:
        ws.append(record.as_row(include_standard))
        ws.cell(row=ws.max_row, column=gtin_col).number_format = "@"
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        width = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 2, 40)
    wb.save(path)


def with_normalized_gtin(record: LabelRecord) -> LabelRecord:
    return replace(record, gtin=normalize_gtin14(record.gtin))
