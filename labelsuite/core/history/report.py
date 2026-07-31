"""LOT별 검사 리포트 내보내기 (xlsx)."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from labelsuite.core.history.db import HistoryDb

_PASS_FILL = PatternFill("solid", start_color="90EE90")
_FAIL_FILL = PatternFill("solid", start_color="FFE4B5")
_BOLD = Font(bold=True)


def export_lot_report(db: HistoryDb, lot: str, out_path: str) -> int:
    """지정 LOT의 검사 이력을 요약+상세 시트로 내보낸다. 반환: 검사 건수."""
    rows = db.query(lot=lot, limit=10_000)
    wb = Workbook()

    summary = wb.active
    summary.title = "요약"
    summary.append(["LOT", lot])
    summary.append(["검사 건수", len(rows)])
    passed_count = sum(1 for r in rows if r.passed)
    summary.append(["합격", passed_count])
    summary.append(["확인 필요", len(rows) - passed_count])
    summary.append(["합격률",
                    f"{passed_count / len(rows) * 100:.1f}%" if rows else "-"])
    for row in summary.iter_rows(max_row=5, max_col=1):
        row[0].font = _BOLD

    detail = wb.create_sheet("검사 상세")
    header = ["일시", "LOT", "REF", "PN", "규격", "소스", "페이지", "판정",
              "이미지 경로", "필드별 검출/기준"]
    detail.append(header)
    for cell in detail[1]:
        cell.font = _BOLD
    for row in rows:
        fields = db.fields_for(row.id)
        field_summary = ", ".join(
            f"{f.field} {f.found}/{f.expected}" for f in fields
            if f.expected is not None)
        detail.append([
            row.ts, row.lot, row.ref, row.pn, row.standard, row.source,
            (row.page + 1) if row.page is not None else "",
            "합격" if row.passed else "확인 필요",
            row.image_path, field_summary])
        verdict_cell = detail.cell(row=detail.max_row, column=8)
        verdict_cell.fill = _PASS_FILL if row.passed else _FAIL_FILL

    for sheet in (summary, detail):
        for col_cells in sheet.columns:
            width = max(len(str(c.value or "")) for c in col_cells)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(width + 2, 50)

    wb.save(out_path)
    return len(rows)
