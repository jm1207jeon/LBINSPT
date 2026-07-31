"""공용 픽스처 — 레거시 입력 엑셀의 실제 컬럼 배치를 재현하는 빌더."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook


def _sparse_row(cells: dict[int, object], width: int) -> list[object]:
    row: list[object] = [None] * width
    for idx, value in cells.items():
        row[idx] = value
    return row


def build_schedule_xlsx(path: Path, rows: list[dict[int, object]]) -> Path:
    """주문일정 체크리스트: 시트 '진행수량', 0-기반 F(5)=LOT G(6)=PN H(7)=PRODUCTS
    J(9)=REF N(13)=국가 X(23)=MFG DATE."""
    wb = Workbook()
    ws = wb.active
    ws.title = "진행수량"
    ws.append(_sparse_row({0: "헤더"}, 30))
    for cells in rows:
        ws.append(_sparse_row(cells, 30))
    wb.save(path)
    return path


def build_product_xlsx(path: Path, rows: list[tuple[str, object]]) -> Path:
    """제품 품목번호 리스트: 시트 '품목번호리스트', A(0)=PN L(11)=GTIN."""
    wb = Workbook()
    ws = wb.active
    ws.title = "품목번호리스트"
    ws.append(_sparse_row({0: "PN", 11: "GTIN"}, 15))
    for pn, gtin in rows:
        ws.append(_sparse_row({0: pn, 11: gtin}, 15))
    wb.save(path)
    return path


def build_bsc_xlsx(path: Path, rows: list[tuple[str, str, object]]) -> Path:
    """BSC FGD 리스트: 시트 '현UPN별', M(12)=REF P(15)=PN AQ(42)=GTIN."""
    wb = Workbook()
    ws = wb.active
    ws.title = "현UPN별"
    ws.append(_sparse_row({12: "REF", 15: "PN", 42: "GTIN"}, 45))
    for ref, pn, gtin in rows:
        ws.append(_sparse_row({12: ref, 15: pn, 42: gtin}, 45))
    wb.save(path)
    return path


def schedule_row(lot="24A1234", pn="HANARO-01", products="HANAROSTENT X",
                 ref="NCN20-080-230", country="독일",
                 mfg=datetime(2024, 5, 10)) -> dict[int, object]:
    return {5: lot, 6: pn, 7: products, 9: ref, 13: country, 23: mfg}


@pytest.fixture
def config_dir(tmp_path):
    """번들 기본값으로 초기화된 임시 설정 디렉터리."""
    from labelsuite.core.config import ensure_defaults

    return ensure_defaults(tmp_path / "config")
