import pytest

from labelsuite.core.config import AppConfig
from labelsuite.core.history.db import HistoryDb
from labelsuite.core.history.report import export_lot_report
from labelsuite.core.inspection import CrossCheckResult, InspectionEngine
from labelsuite.core.ocr.textract_client import OcrWord
from labelsuite.core.schema import LabelRecord
from labelsuite.core.standards import load_standards

RECORD = LabelRecord("25090776", "HANAROSTENT X", "HANARO-01", "NCN20-080-230",
                     "2024-05-10", "2027-05-09", "08806173612345", standard="MDR")


@pytest.fixture
def engine(tmp_path):
    return InspectionEngine(load_standards(AppConfig(tmp_path / "cfg")))


@pytest.fixture
def outcome(engine):
    return engine.inspect(
        RECORD, "MDR", [OcrWord("x-25090776-1", (0, 0, 1, 1), 95)],
        barcode_checks=[CrossCheckResult("DataMatrix", "GTIN",
                                         "08806173612345", "08806173612345", True)])


@pytest.fixture
def db(tmp_path):
    database = HistoryDb(tmp_path / "history.sqlite3")
    yield database
    database.close()


class TestHistoryDb:
    def test_record_and_query(self, db, outcome):
        db.record_inspection(outcome, "/tmp/img.jpg", source="pdf",
                             pdf_path="/tmp/a.pdf", page=0)
        rows = db.query()
        assert len(rows) == 1
        assert rows[0].lot == "25090776"
        assert rows[0].standard == "MDR"
        assert rows[0].passed is False  # 카운트 미달이므로 확인 필요

        fields = db.fields_for(rows[0].id)
        assert any(f.field == "LOT" and f.found == 1 for f in fields)

    def test_query_filters(self, db, outcome):
        db.record_inspection(outcome, "", source="pdf")
        assert db.query(lot="2509") != []
        assert db.query(lot="NOPE") == []
        assert db.query(passed=True) == []
        assert len(db.query(passed=False)) == 1

    def test_counter_persists_across_connections(self, tmp_path):
        """레거시는 실행마다 카운터가 1로 리셋돼 파일을 덮어썼다."""
        path = tmp_path / "history.sqlite3"
        first = HistoryDb(path)
        assert first.next_file_counter() == 1
        assert first.next_file_counter() == 2
        first.close()
        second = HistoryDb(path)
        assert second.next_file_counter() == 3
        second.close()

    def test_cascade_delete(self, db, outcome):
        inspection_id = db.record_inspection(outcome, "", source="camera")
        db.delete(inspection_id)
        assert db.query() == []
        assert db.fields_for(inspection_id) == []

    def test_lots_listing(self, db, outcome):
        db.record_inspection(outcome, "", source="pdf")
        assert db.lots() == ["25090776"]


class TestReport:
    def test_export_lot_report(self, db, outcome, tmp_path):
        db.record_inspection(outcome, "/tmp/img.jpg", source="pdf", page=0)
        out = tmp_path / "report.xlsx"
        count = export_lot_report(db, "25090776", str(out))
        assert count == 1
        assert out.exists()

        from openpyxl import load_workbook

        wb = load_workbook(out)
        assert set(wb.sheetnames) == {"요약", "검사 상세"}
        assert wb["요약"]["B1"].value == "25090776"
        assert wb["검사 상세"].max_row == 2
