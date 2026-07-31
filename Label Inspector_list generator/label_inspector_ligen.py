import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime, timedelta
import os
from collections import defaultdict
import re
import json
import pickle

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

class LabelInspectorLiGen:
    def __init__(self, root):
        # Initialize drag and drop if available
        if HAS_DND:
            self.root = TkinterDnD.Tk() if not isinstance(root, TkinterDnD.Tk) else root
        else:
            self.root = root
        self.root.title("Label Inspector_LiGen")
        self.root.geometry("1200x800")
        
        # File storage
        self.files = {
            'schedule': None,  # 주문일정 체크리스트
            'product': None,   # 제품 품목 번호 리스트
            'bsc': None        # BSC FGD 리스트
        }
        
        # Data storage
        self.schedule_data = None
        self.product_data = None
        self.bsc_data = None
        self.available_dates = []
        self.selected_dates = []
        self.generated_data = None
        
        # Config file path
        self.config_file = os.path.join(os.path.dirname(__file__), 'app_config.json')
        
        # Load saved configuration
        self.load_config()
        
        self.setup_ui()
        
        # UI 생성 후 설정 로드
        self.load_saved_files()
        
    def load_saved_files(self):
        """UI 생성 후 저장된 파일들 로드"""
        try:
            if hasattr(self, 'config_to_load'):
                for key, filepath in self.config_to_load.items():
                    if filepath and os.path.exists(filepath):
                        self.files[key] = filepath
                        self.load_file(filepath, key, update_ui=True)
        except Exception as e:
            print(f"저장된 파일 로드 중 오류: {e}")
        
    def load_config(self):
        """저장된 설정 파일 로드"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                # 파일 경로 복원 - UI 생성 후에만 실행
                self.config_to_load = config.get('files', {})
                        
        except Exception as e:
            print(f"설정 로드 중 오류: {e}")
            self.config_to_load = {}
            
    def save_config(self):
        """현재 설정을 파일에 저장"""
        try:
            config = {
                'files': self.files
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"설정 저장 중 오류: {e}")
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Left panel for buttons
        left_frame = ttk.Frame(main_frame)
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.N), padx=(0, 20))
        
        # File upload buttons
        self.create_file_buttons(left_frame)
        
        # Date selection area
        date_frame = ttk.LabelFrame(left_frame, text="날짜 선택", padding="10")
        date_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(20, 10))
        
        self.date_tree = ttk.Treeview(date_frame, height=10)
        self.date_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbar for date tree
        date_scrollbar = ttk.Scrollbar(date_frame, orient="vertical", command=self.date_tree.yview)
        date_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.date_tree.configure(yscrollcommand=date_scrollbar.set)
        
        # Bind tree events
        self.date_tree.bind('<Button-1>', self.on_date_tree_click)
        
        # Action buttons
        action_frame = ttk.Frame(left_frame)
        action_frame.grid(row=5, column=0, columnspan=2, pady=(10, 0))
        
        self.generate_btn = ttk.Button(action_frame, text="리스트 생성", command=self.generate_list)
        self.generate_btn.grid(row=0, column=0, padx=(0, 10))
        
        self.download_btn = ttk.Button(action_frame, text="다운로드", command=self.download_file, state='disabled')
        self.download_btn.grid(row=0, column=1)
        
        # Right panel for preview - 창 하단까지 확장
        right_frame = ttk.LabelFrame(main_frame, text="미리보기", padding="10")
        right_frame.grid(row=0, column=1, rowspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Preview treeview with smaller font
        style = ttk.Style()
        style.configure("Small.Treeview", font=('TkDefaultFont', 8))
        style.configure("Small.Treeview.Heading", font=('TkDefaultFont', 9, 'bold'))
        
        self.preview_tree = ttk.Treeview(right_frame, columns=('LOT', 'PRODUCTS', 'PN', 'REF', 'MFG_DATE', 'EXP_DATE', 'GTIN'), show='headings', style="Small.Treeview")
        
        # Configure columns with auto-sizing
        self.column_configs = {
            'LOT': {'text': 'LOT', 'width': 80},
            'PRODUCTS': {'text': 'PRODUCTS', 'width': 150},
            'PN': {'text': 'PN', 'width': 100},
            'REF': {'text': 'REF', 'width': 120},
            'MFG_DATE': {'text': 'MFG DATE', 'width': 100},
            'EXP_DATE': {'text': 'EXP DATE', 'width': 100},
            'GTIN': {'text': 'GTIN', 'width': 130}
        }
        
        for i, (col_key, config) in enumerate(self.column_configs.items()):
            self.preview_tree.heading(f'#{i+1}', text=config['text'])
            self.preview_tree.column(f'#{i+1}', width=config['width'], minwidth=50)
        
        self.preview_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Preview scrollbars
        preview_v_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=self.preview_tree.yview)
        preview_v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.preview_tree.configure(yscrollcommand=preview_v_scrollbar.set)
        
        preview_h_scrollbar = ttk.Scrollbar(right_frame, orient="horizontal", command=self.preview_tree.xview)
        preview_h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        self.preview_tree.configure(xscrollcommand=preview_h_scrollbar.set)
        
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(0, weight=1)
        
    def create_file_buttons(self, parent):
        button_names = [
            ("주문일정 체크리스트", "schedule"),
            ("제품 품목 번호 리스트", "product"),
            ("BSC FGD 리스트", "bsc")
        ]
        
        self.file_labels = {}
        
        for i, (name, key) in enumerate(button_names):
            # Button
            btn = ttk.Button(parent, text=name, command=lambda k=key: self.select_file(k))
            btn.grid(row=i, column=0, sticky=(tk.W, tk.E), pady=5)
            
            # Configure drag and drop if available
            if HAS_DND:
                btn.drop_target_register(DND_FILES)
                btn.dnd_bind('<<Drop>>', lambda e, k=key: self.drop_file(e, k))
            
            # File name label
            label = ttk.Label(parent, text="파일이 선택되지 않음", foreground="gray")
            label.grid(row=i, column=1, sticky=(tk.W), padx=(10, 0))
            self.file_labels[key] = label
            
        parent.columnconfigure(0, weight=0)
        parent.columnconfigure(1, weight=1)
        
    def select_file(self, file_type):
        filename = filedialog.askopenfilename(
            title=f"{self.get_file_type_name(file_type)} 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if filename:
            self.load_file(filename, file_type)
            
    def drop_file(self, event, file_type):
        files = self.root.tk.splitlist(event.data)
        if files:
            self.load_file(files[0], file_type)
            
    def get_file_type_name(self, file_type):
        names = {
            'schedule': '주문일정 체크리스트',
            'product': '제품 품목 번호 리스트',
            'bsc': 'BSC FGD 리스트'
        }
        return names.get(file_type, '')
        
    def load_file(self, filename, file_type, update_ui=False):
        try:
            self.files[file_type] = filename
            file_name = os.path.basename(filename)
            if hasattr(self, 'file_labels'):
                self.file_labels[file_type].config(text=file_name, foreground="black")
            
            # Load data based on file type
            if file_type == 'schedule':
                self.schedule_data = pd.read_excel(filename, sheet_name='진행수량')
                self.extract_dates()
            elif file_type == 'product':
                self.product_data = pd.read_excel(filename, sheet_name='품목번호리스트')
            elif file_type == 'bsc':
                self.bsc_data = pd.read_excel(filename, sheet_name='현UPN별')
            
            # 설정 저장
            self.save_config()
            
            if not update_ui:
                messagebox.showinfo("성공", f"{self.get_file_type_name(file_type)} 파일이 로드되었습니다.")
            
        except Exception as e:
            if not update_ui:
                messagebox.showerror("오류", f"파일 로드 중 오류가 발생했습니다: {str(e)}")
            
    def extract_dates(self):
        if self.schedule_data is None:
            return
            
        try:
            # X열 (24번째 열, 0-based index 23)에서 날짜 추출
            date_column = self.schedule_data.iloc[:, 23]  # X열
            
            # 날짜 값들 추출 및 정리
            dates = []
            for date_val in date_column.dropna():
                if pd.notna(date_val):
                    try:
                        if isinstance(date_val, str):
                            # 문자열인 경우 파싱 시도
                            parsed_date = pd.to_datetime(date_val)
                        else:
                            # 이미 datetime 객체인 경우
                            parsed_date = pd.to_datetime(date_val)
                        dates.append(parsed_date.date())
                    except:
                        continue
                        
            # 중복 제거 및 정렬
            self.available_dates = sorted(list(set(dates)))
            self.populate_date_tree()
            
        except Exception as e:
            messagebox.showerror("오류", f"날짜 추출 중 오류가 발생했습니다: {str(e)}")
            
    def populate_date_tree(self):
        # Clear existing items
        for item in self.date_tree.get_children():
            self.date_tree.delete(item)
            
        if not self.available_dates:
            return
            
        # Group dates by year and month
        date_hierarchy = defaultdict(lambda: defaultdict(list))
        
        for date in self.available_dates:
            year = date.year
            month = date.month
            date_hierarchy[year][month].append(date)
            
        # Populate tree
        for year in sorted(date_hierarchy.keys()):
            year_item = self.date_tree.insert('', 'end', text=f'{year}년', values=('year', year), tags=('year',))
            
            for month in sorted(date_hierarchy[year].keys()):
                month_item = self.date_tree.insert(year_item, 'end', text=f'{month}월', values=('month', year, month), tags=('month',))
                
                for date in sorted(date_hierarchy[year][month]):
                    day_item = self.date_tree.insert(month_item, 'end', text=f'{date.day}일 ({date.strftime("%Y-%m-%d")})', 
                                                   values=('day', date.strftime('%Y-%m-%d')), tags=('day',))
                                                   
    def on_date_tree_click(self, event):
        item = self.date_tree.selection()[0] if self.date_tree.selection() else None
        if not item:
            return
            
        values = self.date_tree.item(item, 'values')
        if not values:
            return
            
        if values[0] == 'day':  # Day item clicked
            date_str = values[1]
            
            # Toggle selection
            if date_str in self.selected_dates:
                self.selected_dates.remove(date_str)
                # Remove checkmark
                current_text = self.date_tree.item(item, 'text')
                if current_text.startswith('✓ '):
                    new_text = current_text[2:]
                    self.date_tree.item(item, text=new_text)
            else:
                self.selected_dates.append(date_str)
                # Add checkmark
                current_text = self.date_tree.item(item, 'text')
                if not current_text.startswith('✓ '):
                    new_text = '✓ ' + current_text
                    self.date_tree.item(item, text=new_text)
                    
    def generate_list(self):
        if not self.validate_inputs():
            return
            
        try:
            # Generate data based on selected dates
            result_data = []
            
            for date_str in self.selected_dates:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                # Find matching rows in schedule data
                date_column = self.schedule_data.iloc[:, 23]  # X열 (MFG DATE)
                
                for idx, row in self.schedule_data.iterrows():
                    try:
                        row_date = pd.to_datetime(row.iloc[23]).date() if pd.notna(row.iloc[23]) else None
                        if row_date == date_obj:
                            lot = row.iloc[5] if pd.notna(row.iloc[5]) else ''  # F열 (LOT)
                            products = row.iloc[7] if pd.notna(row.iloc[7]) else ''  # H열 (PRODUCTS)
                            pn = row.iloc[6] if pd.notna(row.iloc[6]) else ''  # G열 (PN)
                            
                            # J열 REF 값 처리 - 실제 REF 데이터만 추출
                            ref_raw = row.iloc[9]  # J열 (REF base)
                            ref_base = ''
                            
                            if pd.notna(ref_raw):
                                ref_str = str(ref_raw).strip()
                                
                                # 날짜 형식 패턴 확인 (YYYY-MM-DD, MM/DD/YYYY 등)
                                date_patterns = [
                                    r'^\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
                                    r'^\d{2}/\d{2}/\d{4}',  # MM/DD/YYYY
                                    r'^\d{4}/\d{2}/\d{2}',  # YYYY/MM/DD
                                    r'^\d{2}-\d{2}-\d{4}',  # MM-DD-YYYY
                                    r'^\d{4}\.\d{2}\.\d{2}', # YYYY.MM.DD
                                    r'^\d{2}\.\d{2}\.\d{4}', # MM.DD.YYYY
                                ]
                                
                                is_date = False
                                for pattern in date_patterns:
                                    if re.match(pattern, ref_str):
                                        is_date = True
                                        break
                                
                                # 시간 정보가 포함된 경우도 확인
                                if '00:00:00' in ref_str or ':' in ref_str:
                                    is_date = True
                                
                                # 숫자만으로 구성되고 40000 이상인 경우 (엑셀 날짜 시리얼)
                                try:
                                    num_val = float(ref_str)
                                    if num_val > 40000 and '.' not in ref_str:
                                        is_date = True
                                except:
                                    pass
                                
                                # 날짜가 아닌 경우만 REF로 사용
                                if not is_date and ref_str not in ['nan', 'NaT', '']:
                                    ref_base = ref_str
                                
                            country = row.iloc[13] if pd.notna(row.iloc[13]) else ''  # N열 (국가)
                            
                            # 날짜 형식을 국가에 따라 다르게 설정
                            if country == '일본':
                                # 일본: YYYY.MM.DD 형식
                                mfg_date = date_obj.strftime('%Y.%m.%d')
                                exp_date = date_obj.replace(year=date_obj.year + 3) - timedelta(days=1)
                                exp_date_str = exp_date.strftime('%Y.%m.%d')
                            else:
                                # 일본 외: YYYY-MM-DD 형식
                                mfg_date = date_obj.strftime('%Y-%m-%d')
                                exp_date = date_obj.replace(year=date_obj.year + 3) - timedelta(days=1)
                                exp_date_str = exp_date.strftime('%Y-%m-%d')
                            
                            # Determine REF
                            ref = ''
                            if country == '일본':
                                # 조건2: 일본인 경우 BSC FGD LIST에서 PN 매칭하여 M열 값 반환
                                if self.bsc_data is not None:
                                    bsc_match = self.bsc_data[self.bsc_data.iloc[:, 15] == pn]  # P열에서 PN 매치
                                    if not bsc_match.empty:
                                        bsc_ref = bsc_match.iloc[0, 12]  # M열
                                        ref = str(bsc_ref) if pd.notna(bsc_ref) else str(ref_base)
                                    else:
                                        ref = str(ref_base)  # BSC에서 매칭되지 않으면 기본값 사용
                                else:
                                    ref = str(ref_base)  # BSC 파일이 없으면 기본값 사용
                            else:
                                # 조건1: 일본이 아닌 모든 경우 S열 값 사용
                                ref = str(ref_base)
                                    
                            # Find GTIN
                            gtin = ''
                            if country == '일본':
                                # 조건2: 일본인 경우 BSC FGD LIST에서 PN 매칭하여 AQ열 값 반환
                                if self.bsc_data is not None:
                                    bsc_match = self.bsc_data[self.bsc_data.iloc[:, 15] == pn]  # P열에서 PN 매치
                                    if not bsc_match.empty:
                                        gtin_val = bsc_match.iloc[0, 42]  # AQ열 (43번째 열, 0-based index 42)
                                        if pd.notna(gtin_val):
                                            gtin = str(gtin_val).strip()
                                        else:
                                            gtin = ''
                            else:
                                # 조건1: 일본이 아닌 모든 경우 제품 품목 번호 리스트에서 추출
                                if self.product_data is not None:
                                    product_match = self.product_data[self.product_data.iloc[:, 0] == pn]  # A열에서 PN 매치
                                    if not product_match.empty:
                                        gtin_val = product_match.iloc[0, 11]  # L열 (GTIN)
                                        if pd.notna(gtin_val):
                                            # GTIN을 텍스트로 처리 (0으로 시작할 수 있음)
                                            if isinstance(gtin_val, (int, float)):
                                                gtin = f"{int(gtin_val):013d}"  # 13자리로 패딩
                                            else:
                                                gtin = str(gtin_val).strip()
                                        else:
                                            gtin = ''
                                    
                            result_data.append({
                                'LOT': lot,
                                'PRODUCTS': products,
                                'PN': pn,
                                'REF': ref,
                                'MFG_DATE': mfg_date,
                                'EXP_DATE': exp_date_str,
                                'GTIN': gtin
                            })
                    except Exception as e:
                        continue
                        
            self.generated_data = pd.DataFrame(result_data)
            self.update_preview()
            self.download_btn.config(state='normal')
            
            messagebox.showinfo("성공", f"{len(result_data)}개의 항목이 생성되었습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"리스트 생성 중 오류가 발생했습니다: {str(e)}")
            
    def validate_inputs(self):
        if not self.files['schedule']:
            messagebox.showerror("오류", "주문일정 체크리스트 파일을 선택해주세요.")
            return False
            
        if not self.selected_dates:
            messagebox.showerror("오류", "날짜를 선택해주세요.")
            return False
            
        return True
        
    def update_preview(self):
        # Clear existing items
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
            
        if self.generated_data is None or self.generated_data.empty:
            return
            
        # Calculate optimal column widths based on content
        max_widths = {}
        for col_key in self.column_configs.keys():
            max_widths[col_key] = len(self.column_configs[col_key]['text']) * 8  # Header width
            
        # Check data widths
        for _, row in self.generated_data.iterrows():
            for col_key in ['LOT', 'PRODUCTS', 'PN', 'REF', 'MFG_DATE', 'EXP_DATE', 'GTIN']:
                content_width = len(str(row[col_key])) * 8
                if content_width > max_widths[col_key]:
                    max_widths[col_key] = min(content_width, 200)  # Max width limit
        
        # Apply calculated widths
        for i, (col_key, width) in enumerate(max_widths.items()):
            self.preview_tree.column(f'#{i+1}', width=max(width, self.column_configs[col_key]['width']))
            
        # Add data to preview
        for _, row in self.generated_data.iterrows():
            self.preview_tree.insert('', 'end', values=(
                row['LOT'], row['PRODUCTS'], row['PN'], row['REF'],
                row['MFG_DATE'], row['EXP_DATE'], row['GTIN']
            ))
            
    def download_file(self):
        if self.generated_data is None or self.generated_data.empty:
            messagebox.showerror("오류", "생성된 데이터가 없습니다.")
            return
            
        # Generate default filename
        date_parts = []
        for date_str in sorted(self.selected_dates):
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            date_parts.append(date_obj.strftime('%y%m%d'))
            
        default_filename = f"Label Inspection List_{','.join(date_parts)}.xlsx"
        
        # Ask for save location
        filename = filedialog.asksaveasfilename(
            title="파일 저장",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                # Create Excel file with proper formatting
                from openpyxl import Workbook
                from openpyxl.styles import NamedStyle
                
                wb = Workbook()
                ws = wb.active
                ws.title = 'Label Inspection List'
                
                # Add headers
                headers = ['LOT', 'PRODUCTS', 'PN', 'REF', 'MFG DATE', 'EXP DATE', 'GTIN']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Add data
                for row_idx, (_, row_data) in enumerate(self.generated_data.iterrows(), 2):
                    ws.cell(row=row_idx, column=1, value=str(row_data['LOT']))
                    ws.cell(row=row_idx, column=2, value=str(row_data['PRODUCTS']))
                    ws.cell(row=row_idx, column=3, value=str(row_data['PN']))
                    ws.cell(row=row_idx, column=4, value=str(row_data['REF']))
                    ws.cell(row=row_idx, column=5, value=str(row_data['MFG_DATE']))
                    ws.cell(row=row_idx, column=6, value=str(row_data['EXP_DATE']))
                    
                    # GTIN을 텍스트로 저장
                    gtin_cell = ws.cell(row=row_idx, column=7, value=str(row_data['GTIN']))
                    gtin_cell.number_format = '@'
                
                wb.save(filename)
                        
                messagebox.showinfo("성공", f"파일이 저장되었습니다: {filename}")
                
            except Exception as e:
                messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {str(e)}")

def main():
    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    app = LabelInspectorLiGen(root)
    root.mainloop()

if __name__ == "__main__":
    main()
